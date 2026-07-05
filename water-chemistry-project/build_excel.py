"""
Builds carbonate_system_data.xlsx: one sheet per figure with the underlying
data AND a native Excel chart (same role as Yosef's HWF.OUTPOT*.xlsx files).
Run after carbonate_system_model.py is importable.
"""
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import LineChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
import carbonate_system_model as m

HEAD_FILL = PatternFill('solid', fgColor='1B3A53')
HEAD_FONT = Font(bold=True, color='FFFFFF')


def new_sheet(wb, title, headers, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEAD_FONT; cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(12, len(h) + 2)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=float(v))
    ws.freeze_panes = 'A2'
    return ws


def add_chart(ws, title, x_title, y_title, n_rows, y_cols, x_col=1,
              anchor='H2', scatter=False):
    """y_cols: list of 1-based column indices to plot on Y."""
    chart = ScatterChart() if scatter else LineChart()
    chart.title = title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.height = 9; chart.width = 16
    chart.style = 2
    xref = Reference(ws, min_col=x_col, min_row=2, max_row=n_rows + 1)
    for col in y_cols:
        yref = Reference(ws, min_col=col, min_row=1, max_row=n_rows + 1)
        if scatter:
            s = Series(yref, xref, title_from_data=True)
            chart.series.append(s)
        else:
            chart.add_data(yref, titles_from_data=True)
    if not scatter:
        chart.set_categories(xref)
    ws.add_chart(chart, anchor)
    return chart


wb = Workbook()

# ---- 1. Speciation (Bjerrum) ------------------------------------------------
cs = m.CarbonateSystem(T_C=25, I=0.01)
pH = np.arange(2, 13.01, 0.1)
H = 10 ** (-pH) / cs.g1
d = 1 + cs.K1c / H + cs.K1c * cs.K2c / H ** 2
rows = list(zip(pH, 1 / d, (cs.K1c / H) / d, (cs.K1c * cs.K2c / H ** 2) / d))
ws = new_sheet(wb, 'Speciation',
               ['pH', 'alpha0 (H2CO3*)', 'alpha1 (HCO3-)', 'alpha2 (CO3-2)'],
               rows, first=True)
add_chart(ws, 'Carbonate speciation vs pH', 'pH', 'mole fraction',
          len(rows), [2, 3, 4])

# ---- 2. Acidification -------------------------------------------------------
ppm = np.logspace(np.log10(150), np.log10(3000), 60)
rows = []
for p in ppm:
    c = m.CarbonateSystem(T_C=15, I=0.01, Alk=2.0e-3, Ca=0.8e-3,
                          pCO2=p * 1e-6, mode='open')
    rows.append((p, c.pH, c.CT * 1e3))
ws = new_sheet(wb, 'Acidification', ['pCO2 (ppm)', 'pH', 'CT (mmol/L)'], rows)
add_chart(ws, 'pH vs pCO2 (acidification)', 'pCO2 (ppm)', 'pH',
          len(rows), [2], scatter=True)

# ---- 3. Calcite saturation --------------------------------------------------
rows = []
for p in ppm:
    c = m.CarbonateSystem(T_C=15, I=0.01, Alk=2.0e-3, Ca=0.8e-3,
                          pCO2=p * 1e-6, mode='open')
    rows.append((p, c.SI, c.Omega))
ws = new_sheet(wb, 'Calcite_SI', ['pCO2 (ppm)', 'SI', 'Omega'], rows)
add_chart(ws, 'Calcite SI vs pCO2', 'pCO2 (ppm)', 'SI',
          len(rows), [2], scatter=True)

# ---- 4. Temperature ---------------------------------------------------------
rows = []
for t in np.linspace(0, 40, 41):
    c = m.CarbonateSystem(T_C=t, I=0.01, Alk=2.0e-3, Ca=0.8e-3,
                          pCO2=400e-6, mode='open')
    rows.append((t, c.pH, c.SI))
ws = new_sheet(wb, 'Temperature', ['T (C)', 'pH', 'SI'], rows)
add_chart(ws, 'Effect of temperature', 'T (C)', 'value', len(rows), [2, 3])

# ---- 5. Ionic strength ------------------------------------------------------
rows = []
for I in np.linspace(1e-4, 0.5, 60):
    c = m.CarbonateSystem(T_C=15, I=I, Alk=2.0e-3, Ca=0.8e-3,
                          pCO2=400e-6, mode='open')
    rows.append((I, -np.log10(c.K1c), -np.log10(c.K2c), c.SI))
ws = new_sheet(wb, 'IonicStrength',
               ['I (mol/L)', "pK1'", "pK2'", 'SI'], rows)
add_chart(ws, 'Effect of ionic strength', 'I (mol/L)', 'value',
          len(rows), [2, 3, 4], scatter=True)

# ---- 6. Kinetics (dynamic) --------------------------------------------------
r = m.simulate_kinetics(n=240)
step = 2   # subsample to ~120 rows
rows = list(zip(r['t'][::step], r['pH'][::step], r['CT'][::step],
                r['Ca'][::step], r['SI'][::step]))
ws = new_sheet(wb, 'Kinetics',
               ['t (h)', 'pH', 'CT (mmol/L)', 'Ca (mmol/L)', 'SI'], rows)
add_chart(ws, 'Dynamic: degassing + calcite precipitation', 't (h)', 'value',
          len(rows), [2, 4], scatter=True)

# ---- README sheet -----------------------------------------------------------
ws = wb.create_sheet('README', 0)
ws.column_dimensions['A'].width = 100
notes = [
    'Aquatic Carbonate System & Water Acidification - simulation output data',
    'Water Chemistry course, BGU, Spring 2026',
    '',
    'Each sheet holds the data behind one report figure, with an embedded chart:',
    '  Speciation    - Bjerrum plot (mole fractions vs pH)',
    '  Acidification - pH & CT vs atmospheric pCO2',
    '  Calcite_SI    - calcite saturation index & Omega vs pCO2',
    '  Temperature   - pH & SI vs temperature',
    '  IonicStrength - apparent pK1, pK2 & SI vs ionic strength',
    '  Kinetics      - dynamic run: CO2 mass transfer + calcite precipitation kinetics',
    '',
    'Generated by build_excel.py from carbonate_system_model.py.',
]
for i, line in enumerate(notes, 1):
    cell = ws.cell(row=i, column=1, value=line)
    if i == 1:
        cell.font = Font(bold=True, size=13)

wb.save('carbonate_system_data.xlsx')
print('Wrote carbonate_system_data.xlsx  (sheets:', wb.sheetnames, ')')
